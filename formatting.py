"""
Result Transforming - Excel Formatting

Builds the enriched mismatch report using ``processing.run_processing`` (same logic as
``processing.py``), then applies presentation rules:

- Column renaming (COLUMN_RENAME_MAP)
- Optional extra columns from ``lc_etof_with_comments.xlsx``
- Header styling, alternating row colors for cost-type groups, column widths, borders,
  freeze panes

By default this runs enrichment first, then formats. Set ``run_enrichment=False`` to only
format an existing workbook (e.g. already-written ``mismatch_enriched.xlsx``).
"""

from __future__ import annotations

import os
from pathlib import Path
from typing import Optional

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from processing import (
    COL_ANOTHER_RC_LANE_VS_SHIPMENT,
    COL_APPLIES_IF,
    COL_ANOTHER_RATE_CARD_CARRIER_USED_CRF,
    COL_ANOTHER_RATE_CARD_CARRIER_USED_INV,
    COL_BEST_MATCH_ANOTHER_RATE_CARD,
    COL_CARRIER_RATE_FILE,
    COL_EXCHANGE_RATE,
    COL_POSSIBLE_CARRIER_EXCHANGE_RATE,
    COL_POSSIBLE_RATE_CARD_VALUE_USED,
    COL_RATE_BY,
    COL_RATE_COST,
    COL_RATE_COST_CALCULATED,
    COL_RATE_COST_FILE,
    COL_ROUNDING_RULE,
    run_processing,
)

ENRICHED_SHEET_NAME = "Enriched"

# Final workbook path (after rename, column filter, reorder, styling)
OUTPUT_FOLDER_NAME = "output"
OUTPUT_XLSX_FILENAME = "mismatch_report.xlsx"

# Dropped from the formatted export (still present in partly_df ``mismatch_enriched.*`` from processing)
FORMATTING_DROP_COLUMNS: frozenset[str] = frozenset(
    {
        COL_EXCHANGE_RATE,
        COL_APPLIES_IF,
        COL_RATE_BY,
        COL_ROUNDING_RULE,
        COL_POSSIBLE_RATE_CARD_VALUE_USED,
        COL_POSSIBLE_CARRIER_EXCHANGE_RATE,
        COL_ANOTHER_RATE_CARD_CARRIER_USED_CRF,
        COL_ANOTHER_RATE_CARD_CARRIER_USED_INV,
        COL_BEST_MATCH_ANOTHER_RATE_CARD,
        COL_RATE_COST,
        COL_RATE_COST_CALCULATED,
        COL_RATE_COST_FILE,
        COL_CARRIER_RATE_FILE,
        # Additional fields excluded from the final report
        "ISD_NUMBER",
        "FID #",
        "CARRIER_REFERENCE",
        "INVOICE_NUMBER",
        "INVOICE_ENTITY",
        "SHIP_COUNTRY_ISD",
        "CUST_COUNTRY_ISD",
        "SERVICE_ISD",
        "SHIP_CITY_ISD",
        "CUST_CITY_ISD",
        "SHIP_AIRPORT_ISD",
        "CARRIER_ACCOUNT_NR_ISD",
        "BUSINESS_SEGMENT",
        "WEIGHT_ETOF",
        "WEIGHT_ISD",
        "LC",
        "Billing account",
        "TRANSPORT_MODE",
        "ORIGINAL_SERVICE",
        "Carrier agreement #",
        "ORIG_FILE_NAME",
        "INV_TYPE",
    }
)

# Renamed ``ETOF_NUMBER`` → ``ETOF``; rows are sorted by this column in the final export
SORT_BY_ETOF_COLUMN = "ETOF"

# Column order after COLUMN_RENAME_MAP is applied (post-rename names only)
OUTPUT_COLUMN_ORDER: list[str] = [
    "ETOF",
    "Delivery Number",
    "Shipment date",
    "BU_NAME",
    "CONT_LOAD",
    "Origin country",
    "Destination country",
    "Service",
    "Origin city",
    "Destination city",
    "SHIP_AIRPORT_ETOF",
    "CUST_AIRPORT",
    "CARRIER_ACCOUNT_NR_ETOF",
    "CARRIER_NAME",
    "SERVICE",
    "SHIP_POST",
    "CUST_POST",
    "CHARGEABLE WEIGHT",
    "Cost type",
    "Invoice currency",
    "Pre-calc. cost (in inv curr)",
    "Invoice statement cost  (in inv curr)",
    "Discrepancy in inv currency  (in inv curr)",
    "Agreement RA",
    "best_lane(s)",
    COL_ANOTHER_RC_LANE_VS_SHIPMENT,
]


# Column renaming mapping (original -> new name) - ALWAYS APPLIED
COLUMN_RENAME_MAP = {
    'ETOF_NUMBER': 'ETOF',
    'SHIPMENT_ID': 'Shipment ID',
    'DELIVERY_NUMBER': 'Delivery Number',
    'SHIP_DATE': 'Shipment date',
    'SHIP_COUNTRY_ETOF': 'Origin country',
    'SHIP_CITY_ETOF': 'Origin city',
    'CUST_COUNTRY_ETOF': 'Destination country',
    'CUST_CITY_ETOF': 'Destination city',
    'SERVICE_ETOF': 'Service',
}


# =============================================================================
# EXTRA COLUMNS ALIAS MAP
# =============================================================================
# This mapping defines what the user can type -> what actual column name to look for
# Format: 'what user can type': 'actual column name in source file'
# Add your own mappings here!
# =============================================================================
EXTRA_COLUMNS_ALIAS_MAP = {
    # Format: 'Display name (from dropdown)': 'Actual column name in source file'
    'Invoice entity': 'INVOICE_ENTITY',
    'Carrier name': 'CARRIER_NAME',
    'Destination postal code': 'CUST_POST',
    'Origin postal code': 'SHIP_POST',
    'Destination airport': 'CUST_AIRPORT',
    'Equipment type': 'CONT_LOAD',
    'Origin airport': 'SHIP_AIRPORT',
    'Business unit name': 'BU_NAME',
    'Transport mode': 'TRANSPORT_MODE',
    'LDM': 'LDM',
    'CBM': 'CBM',
    'Weight': 'WEIGHT',
    'DANGEROUS Goods': 'DANGEROUS_GOODS',
    'Charge weight': 'CHARGE_WEIGHT',
    'House bill': 'HOUSE_BILL',
    'Master bill': 'MASTER_BILL',
    'Roundtrip': 'ROUNDTRIP',
}


def get_column_aliases(col_name):
    """
    Get all possible column names to search for based on user input.
    
    Checks EXTRA_COLUMNS_ALIAS_MAP first, then returns as-is if not found.
    
    Args:
        col_name: Column name provided by user
    
    Returns:
        List of possible column names to search for in source file
    """
    col_stripped = col_name.strip()
    col_lower = col_stripped.lower()
    
    # Check if user input matches any alias in EXTRA_COLUMNS_ALIAS_MAP
    for alias, actual_col in EXTRA_COLUMNS_ALIAS_MAP.items():
        if col_lower == alias.lower():
            # Return both the actual column and the alias (in case source uses alias)
            return [actual_col, col_stripped]
    
    # If not in alias map, return as-is
    return [col_stripped]


def rename_columns(df):
    """
    Rename columns according to the COLUMN_RENAME_MAP.
    
    Args:
        df: DataFrame to rename columns in
    
    Returns:
        DataFrame with renamed columns
    """
    # Create a case-insensitive mapping
    rename_map = {}
    for col in df.columns:
        col_upper = col.upper().strip()
        for old_name, new_name in COLUMN_RENAME_MAP.items():
            if col_upper == old_name.upper():
                rename_map[col] = new_name
                break
    
    if rename_map:
        print(f"      Renaming columns: {rename_map}")
        df = df.rename(columns=rename_map)
    
    return df


def drop_formatting_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Remove columns not needed in the final export."""
    present = [c for c in df.columns if c in FORMATTING_DROP_COLUMNS]
    if not present:
        return df
    print(f"      Dropping columns: {present}")
    return df.drop(columns=present)


def reorder_output_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Apply ``OUTPUT_COLUMN_ORDER``, then any remaining columns alphabetically."""
    ordered: list[str] = []
    for c in OUTPUT_COLUMN_ORDER:
        if c in df.columns:
            ordered.append(c)
    rest = sorted(c for c in df.columns if c not in ordered)
    return df[ordered + rest]


def sort_rows_by_etof(df: pd.DataFrame) -> pd.DataFrame:
    """Sort rows by ``ETOF`` (string order, case-insensitive; missing last)."""
    if df.empty or SORT_BY_ETOF_COLUMN not in df.columns:
        return df
    return df.sort_values(
        by=SORT_BY_ETOF_COLUMN,
        key=lambda s: s.astype(str).str.upper(),
        na_position="last",
    ).reset_index(drop=True)


def get_partly_df_folder() -> Path:
    """Get the path to the partly_df folder."""
    return Path(__file__).parent / "partly_df"


def get_output_folder() -> Path:
    """Folder for the final formatted workbook (created if missing)."""
    p = Path(__file__).parent / OUTPUT_FOLDER_NAME
    p.mkdir(parents=True, exist_ok=True)
    return p


def build_cost_type_groups_from_dataframe(
    df: pd.DataFrame, sheet_name: str = ENRICHED_SHEET_NAME
) -> dict[str, list[tuple[int, int, int]]]:
    """
    Build ``cost_type_groups`` for :func:`apply_formatting`: contiguous rows with the same
    ``Cost type`` get alternating fill (color_index 0 vs 1). Excel rows are 1-based; data
    starts at row 2.
    """
    col = "Cost type"
    if df.empty or col not in df.columns:
        return {}
    groups: list[tuple[int, int, int]] = []
    start_idx = 0
    color_idx = 0

    def _ct(i: int) -> str:
        v = df.iloc[i].get(col)
        return str(v).strip() if v is not None and not (isinstance(v, float) and pd.isna(v)) else ""

    for i in range(1, len(df)):
        if _ct(i) != _ct(i - 1):
            end_idx = i - 1
            groups.append((start_idx + 2, end_idx + 2, color_idx))
            color_idx = 1 - color_idx
            start_idx = i
    end_idx = len(df) - 1
    groups.append((start_idx + 2, end_idx + 2, color_idx))
    return {sheet_name: groups}


def find_etof_column(df):
    """Find the ETOF number column in a DataFrame."""
    etof_patterns = ['etof', 'etof_number', 'etof number', 'etof#', 'etof #']
    for col in df.columns:
        col_lower = col.lower().strip()
        for pattern in etof_patterns:
            if pattern in col_lower:
                return col
    return None


def add_columns_from_source(result_df, columns_to_add, sheet_name=None):
    """
    Add specified columns from lc_etof_with_comments.xlsx to the result DataFrame.
    
    Args:
        result_df: DataFrame to add columns to
        columns_to_add: List of column names to extract and add
        sheet_name: Optional sheet name for logging
    
    Returns:
        DataFrame with added columns
    """
    if not columns_to_add:
        return result_df
    
    partly_df = get_partly_df_folder()
    source_file = partly_df / "lc_etof_with_comments.xlsx"
    
    if not source_file.exists():
        print(f"      [WARNING] Source file not found: {source_file}")
        return result_df
    
    # Load source file
    try:
        source_df = pd.read_excel(source_file)
        print(f"      Loaded source file: {len(source_df)} rows")
    except Exception as e:
        print(f"      [WARNING] Error loading source file: {e}")
        return result_df
    
    # Find ETOF column in both DataFrames
    result_etof_col = find_etof_column(result_df)
    source_etof_col = find_etof_column(source_df)
    
    if result_etof_col is None:
        print(f"      [WARNING] ETOF column not found in result DataFrame")
        return result_df
    
    if source_etof_col is None:
        print(f"      [WARNING] ETOF column not found in source file")
        return result_df
    
    print(f"      Matching on: result[{result_etof_col}] <-> source[{source_etof_col}]")
    
    # Find which columns exist in source (using alias mapping)
    columns_found = []
    columns_not_found = []
    
    for col in columns_to_add:
        col_stripped = col.strip()
        found = False
        
        # Get possible aliases from EXTRA_COLUMNS_ALIAS_MAP
        aliases = get_column_aliases(col_stripped)
        
        # Try to find any of the aliases in the source DataFrame
        for alias in aliases:
            matching_cols = [c for c in source_df.columns if c.lower().strip() == alias.lower()]
            if matching_cols:
                columns_found.append(matching_cols[0])  # Use the actual column name from source
                if alias != col_stripped:
                    print(f"      Column '{col_stripped}' mapped to '{matching_cols[0]}'")
                found = True
                break
        
        if not found:
            columns_not_found.append(col_stripped)
    
    if columns_not_found:
        print(f"      [WARNING] Columns not found in source: {columns_not_found}")
        print(f"      Tip: Add mappings to EXTRA_COLUMNS_ALIAS_MAP in formatting.py")
    
    if not columns_found:
        print(f"      [WARNING] No requested columns found in source file")
        return result_df
    
    print(f"      Adding columns: {columns_found}")
    
    # Create a subset of source with ETOF and requested columns
    source_subset = source_df[[source_etof_col] + columns_found].copy()
    
    # Remove duplicates based on ETOF (keep first)
    source_subset = source_subset.drop_duplicates(subset=[source_etof_col], keep='first')
    
    # Rename source ETOF column to match result for merging
    source_subset = source_subset.rename(columns={source_etof_col: result_etof_col})
    
    # Merge with result
    result_with_cols = result_df.merge(
        source_subset,
        on=result_etof_col,
        how='left',
        suffixes=('', '_added')
    )
    
    # Rename the added columns to friendly names (reverse of EXTRA_COLUMNS_ALIAS_MAP)
    # EXTRA_COLUMNS_ALIAS_MAP: 'friendly name' -> 'original name'
    # We need: 'original name' -> 'friendly name'
    reverse_alias_map = {v: k for k, v in EXTRA_COLUMNS_ALIAS_MAP.items()}
    rename_added = {}
    for col in columns_found:
        if col in reverse_alias_map:
            rename_added[col] = reverse_alias_map[col]
    
    if rename_added:
        print(f"      Renaming added columns: {rename_added}")
        result_with_cols = result_with_cols.rename(columns=rename_added)
    
    # Reorder columns: put added columns BEFORE "Pre-calc. cost"
    precalc_col = None
    for col in result_with_cols.columns:
        if 'pre-calc' in col.lower() or 'precalc' in col.lower():
            precalc_col = col
            break
    
    if precalc_col:
        # Get the position of Pre-calc. cost
        cols = list(result_with_cols.columns)
        precalc_idx = cols.index(precalc_col)
        
        # Get the names of added columns (after renaming)
        added_col_names = [rename_added.get(c, c) for c in columns_found]
        
        # Remove added columns from their current positions
        other_cols = [c for c in cols if c not in added_col_names]
        
        # Find where precalc_col is now in other_cols
        precalc_idx_new = other_cols.index(precalc_col)
        
        # Insert added columns before precalc
        new_order = other_cols[:precalc_idx_new] + added_col_names + other_cols[precalc_idx_new:]
        
        result_with_cols = result_with_cols[new_order]
        print(f"      Columns repositioned before '{precalc_col}'")
    
    print(f"      Columns added successfully")
    return result_with_cols


def apply_formatting(wb, cost_type_groups=None):
    """
    Apply formatting to the workbook.
    
    Args:
        wb: openpyxl Workbook object
        cost_type_groups: dict {sheet_name: list of (start_row, end_row, color_index), ...}
                         where color_index is 0 or 1 for alternating colors
    """
    if cost_type_groups is None:
        cost_type_groups = {}
    
    # Define styles
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    pivot_header_fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Two alternating colors for cost type groups
    cost_color_1 = PatternFill(start_color='DAEEF3', end_color='DAEEF3', fill_type='solid')  # Light blue
    cost_color_2 = None  # White (no fill)
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        # Determine if this is a pivot sheet
        is_pivot = 'Pivot' in sheet_name
        
        # Format header row
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = pivot_header_fill if is_pivot else header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Get cost type groups for this sheet
        groups = cost_type_groups.get(sheet_name, [])
        
        # Create a row -> color mapping from groups
        row_colors = {}
        for start_row, end_row, color_idx in groups:
            for r in range(start_row, end_row + 1):
                row_colors[r] = color_idx
        
        # Format data rows
        for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(vertical='center', wrap_text=True)
                
                # Apply color based on cost type group (for data sheets)
                if not is_pivot and row_idx in row_colors:
                    if row_colors[row_idx] == 0:
                        cell.fill = cost_color_1
                    # else: leave white (no fill needed)
        
        # Auto-adjust column widths
        for column_cells in ws.columns:
            max_length = 0
            column = column_cells[0].column_letter
            
            for cell in column_cells:
                try:
                    if cell.value:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                except:
                    pass
            
            # Set width with limits
            adjusted_width = min(max(max_length + 2, 10), 50)
            ws.column_dimensions[column].width = adjusted_width
        
        # Freeze the header row
        ws.freeze_panes = 'A2'


def format_result_file(
    file_path: Optional[os.PathLike[str] | str] = None,
    cost_type_groups: Optional[dict[str, list[tuple[int, int, int]]]] = None,
    extra_columns: Optional[list[str]] = None,
    *,
    run_enrichment: bool = True,
    partly_df: Optional[os.PathLike[str] | str] = None,
    mismatch_json: Optional[os.PathLike[str] | str] = None,
    matched_json: Optional[os.PathLike[str] | str] = None,
    output_json: Optional[os.PathLike[str] | str] = None,
    output_xlsx: Optional[os.PathLike[str] | str] = None,
    formatted_output_path: Optional[os.PathLike[str] | str] = None,
    auto_cost_type_stripes: bool = True,
) -> Path:
    """
    Produce the enriched workbook (``processing.run_processing``) when ``run_enrichment``
    is True, then rename columns, optionally merge extra columns, drop technical columns,
    reorder columns, write the **final** workbook under ``output/``, and apply openpyxl
    styling.

    Args:
        file_path: Source workbook when ``run_enrichment`` is False. When enrichment runs,
            the processing output path is ``output_xlsx`` if set, otherwise ``file_path``, otherwise
            ``partly_df/mismatch_enriched.xlsx`` (raw enrichment; unchanged after this run).
        formatted_output_path: Final formatted file (default: ``output/mismatch_report.xlsx``).
        cost_type_groups: Per-sheet row ranges for alternating fills. If None and
            ``auto_cost_type_stripes`` is True, groups are derived from the ``Cost type`` column.
        extra_columns: Display names to pull from ``lc_etof_with_comments.xlsx`` (see
            EXTRA_COLUMNS_ALIAS_MAP).
        run_enrichment: If True, call :func:`processing.run_processing` first (same inputs
            as the processing CLI).
        partly_df, mismatch_json, matched_json, output_json, output_xlsx: Passed through
            to :func:`processing.run_processing` when ``run_enrichment`` is True.

    Returns:
        Path to the **formatted** workbook under ``output/``
    """
    partly_df_p = Path(partly_df) if partly_df else get_partly_df_folder()

    if run_enrichment:
        out_json: Optional[str] = None
        out_xlsx: Optional[str] = None
        if output_json is not None:
            out_json = os.fspath(output_json)
        if output_xlsx is not None:
            out_xlsx = os.fspath(output_xlsx)
        elif file_path is not None:
            out_xlsx = os.fspath(file_path)
        else:
            out_xlsx = str(partly_df_p / "mismatch_enriched.xlsx")

        print("\n   Running processing.enrichment (run_processing)…")
        _, written_xlsx = run_processing(
            partly_df=os.fspath(partly_df_p),
            mismatch_json=os.fspath(mismatch_json) if mismatch_json else None,
            matched_json=os.fspath(matched_json) if matched_json else None,
            output_json=out_json,
            output_xlsx=out_xlsx,
        )
        file_path = Path(written_xlsx)
    else:
        if file_path is None:
            file_path = partly_df_p / "mismatch_enriched.xlsx"
        else:
            file_path = Path(file_path)

    if not file_path.exists():
        raise FileNotFoundError(f"File not found: {file_path}")

    out_final = (
        Path(formatted_output_path)
        if formatted_output_path is not None
        else get_output_folder() / OUTPUT_XLSX_FILENAME
    )
    out_final.parent.mkdir(parents=True, exist_ok=True)

    # Step 1: Load enriched source, rename columns, optional extra columns, drop, reorder
    print(f"\n   Transforming result file: {file_path}")
    print(f"   Final export will be written to: {out_final}")

    xlsx = pd.ExcelFile(file_path)
    transformed_sheets: dict[str, pd.DataFrame] = {}

    for sheet_name in xlsx.sheet_names:
        df = pd.read_excel(xlsx, sheet_name=sheet_name)

        # Only process data sheets (not pivot sheets)
        if "Pivot" not in sheet_name:
            print(f"\n   Processing sheet: {sheet_name}")

            # Rename columns
            df = rename_columns(df)

            # Add extra columns if specified
            if extra_columns:
                print(f"      Adding extra columns: {extra_columns}")
                df = add_columns_from_source(df, extra_columns, sheet_name)

            df = drop_formatting_columns(df)
            df = reorder_output_columns(df)
            df = sort_rows_by_etof(df)

        transformed_sheets[sheet_name] = df

    effective_groups: dict[str, list[tuple[int, int, int]]] = dict(cost_type_groups or {})
    if auto_cost_type_stripes and not cost_type_groups:
        for sn, sdf in transformed_sheets.items():
            if "Pivot" not in sn:
                effective_groups.update(build_cost_type_groups_from_dataframe(sdf, sn))

    # Save formatted workbook to output/ (not the partly_df enrichment file)
    with pd.ExcelWriter(out_final, engine="openpyxl") as writer:
        for sheet_name, df in transformed_sheets.items():
            df.to_excel(writer, sheet_name=sheet_name, index=False)

    print(f"\n   Saved formatted workbook: {out_final}")

    # Step 2: Apply formatting
    print(f"\n   Applying formatting to: {out_final}")

    wb = load_workbook(out_final)
    apply_formatting(wb, effective_groups)
    wb.save(out_final)

    print("   Formatting applied successfully")
    return out_final


def main(
    file_path: Optional[os.PathLike[str] | str] = None,
    cost_type_groups: Optional[dict[str, list[tuple[int, int, int]]]] = None,
    extra_columns: Optional[list[str]] = None,
    *,
    run_enrichment: bool = True,
    partly_df: Optional[os.PathLike[str] | str] = None,
    mismatch_json: Optional[os.PathLike[str] | str] = None,
    matched_json: Optional[os.PathLike[str] | str] = None,
    output_json: Optional[os.PathLike[str] | str] = None,
    output_xlsx: Optional[os.PathLike[str] | str] = None,
    formatted_output_path: Optional[os.PathLike[str] | str] = None,
    auto_cost_type_stripes: bool = True,
) -> Path:
    """
    Default entry: run enrichment from ``partly_df`` (same as ``processing.py``), then
    apply this module's Excel structure (renames, optional extra columns, column filter,
    order, styling). Writes the final file to ``output/mismatch_report.xlsx``.

    Set ``run_enrichment=False`` to format an existing ``mismatch_enriched.xlsx`` without
    re-running enrichment.
    """
    return format_result_file(
        file_path=file_path,
        cost_type_groups=cost_type_groups,
        extra_columns=extra_columns,
        run_enrichment=run_enrichment,
        partly_df=partly_df,
        mismatch_json=mismatch_json,
        matched_json=matched_json,
        output_json=output_json,
        output_xlsx=output_xlsx,
        formatted_output_path=formatted_output_path,
        auto_cost_type_stripes=auto_cost_type_stripes,
    )


if __name__ == "__main__":
    main()
