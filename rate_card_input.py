import json
import math
import re
import pandas as pd
import openpyxl
import os


def extract_rate_agreement_id_from_filename(file_path: str) -> str:
    """
    From names like 'Advanced Export - RA20250826013 v.9 - DHL_FY25_ME (1).xlsx'
    return 'RA20250826013'. If no RA######## pattern, use a safe stem fallback.
    """
    base = os.path.basename(file_path)
    m = re.search(r"(RA\d+)", base, re.IGNORECASE)
    if m:
        return m.group(1).upper()
    stem = os.path.splitext(base)[0]
    safe = re.sub(r"[^\w\-]+", "_", stem).strip("_")
    return safe[:80] if safe else "rate_card"


def default_filtered_rate_card_basename(file_path: str) -> str:
    """e.g. Filtered_Rate_Card_with_Conditions_RA20250826013"""
    return f"Filtered_Rate_Card_with_Conditions_{extract_rate_agreement_id_from_filename(file_path)}"


def extract_ra_id_from_carrier_agreement(value):
    """
    From ETOF values like 'RA20250326009 (v.13) - On Hold', return 'RA20250326009'
    (RA + digits before any '(v...)' clause). Returns None if not found.
    """
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return None
    s = str(value).strip()
    if not s or s.lower() in ("nan", "none", ""):
        return None
    head = re.split(r"\s*\(\s*[vV]", s, maxsplit=1)[0].strip()
    m = re.search(r"(RA\d+)", head, re.IGNORECASE)
    if m:
        return m.group(1).upper()
    m = re.search(r"(RA\d+)", s, re.IGNORECASE)
    return m.group(1).upper() if m else None


def find_rate_card_xlsx_basename_for_ra_id(ra_id: str, input_folder: str = "input"):
    """
    Return the basename of an .xlsx in ``input_folder`` whose filename contains ``ra_id``
    (e.g. RA20250326009). Prefers the first match in sorted order.
    """
    if not ra_id or not os.path.isdir(input_folder):
        return None
    ra_u = ra_id.upper()
    matches = [
        n for n in os.listdir(input_folder)
        if n.lower().endswith(".xlsx") and ra_u in n.upper()
    ]
    return sorted(matches)[0] if matches else None


def filtered_rate_card_json_path_for_ra_id(ra_id: str, partly_df_folder: str) -> str:
    """partly_df/Filtered_Rate_Card_with_Conditions_<RA id>.json"""
    return os.path.join(partly_df_folder, f"Filtered_Rate_Card_with_Conditions_{ra_id}.json")


def process_rate_card(file_path):
    """
    Process a Rate Card Excel file from the input folder.
    
    Args:
        file_path (str): Path to the file relative to the "input/" folder (e.g., "rate_card.xlsx")
    
    Returns:
        tuple: (dataframe, list of column names, conditions_list)
            - dataframe: Processed pandas DataFrame (filtered to black font columns)
            - list: List of column names in the processed dataframe
            - list: Cleaned condition text per column (same order as column names; may be "")
    """
    # Construct full path from input folder
    input_folder = "input"
    full_path = os.path.join(input_folder, file_path)
    
    # Read the Excel file
    df_rate_card = pd.read_excel(full_path, sheet_name="Rate card", skiprows=2)
    
    # Find first column index (where data actually starts)
    first_column_index = None
    if df_rate_card is not None:
        for i, col in enumerate(df_rate_card.columns):
            if "nan" not in str(df_rate_card.iloc[0, i]).lower():
                first_column_index = i
                break
    
    if first_column_index is not None:
        df_rate_card = df_rate_card.iloc[:, :first_column_index]
    
    # Drop rows where the first column is NaN
    if df_rate_card is not None:
        df_rate_card.dropna(subset=[df_rate_card.columns[0]], inplace=True)
    
    # Set column names from first row
    new_columns = df_rate_card.iloc[0].tolist()
    df_rate_card.columns = new_columns
    df_rate_card = df_rate_card.iloc[1:]
    
    # Load the workbook to extract conditions and check font colors
    workbook = openpyxl.load_workbook(full_path, data_only=True)
    sheet = workbook["Rate card"]
    
    # Find the header row that contains "Currency"
    first_data_row_index = None
    currency_index = None
    
    for row_index in range(1, min(151, sheet.max_row + 1)):
        row = sheet[row_index]
        row_values = [cell.value for cell in row]
        if "Currency" in row_values:
            currency_index = row_values.index("Currency")
            first_data_row_index = row_index
            break
    
    black_font_values = []
    truncated_data_values = None
    column_notes_by_excel_col = {}  # 1-based Excel column index -> raw note text
    valid_indices = None

    if first_data_row_index is not None and currency_index is not None:
        # Access the data in this row
        first_data_row = sheet[first_data_row_index]
        first_data_values = [cell.value for cell in first_data_row]
        truncated_data_values = first_data_values[:currency_index]

        # Extract conditional rules: comment on header cell, else cell directly above header,
        # else row 2 (legacy). Keyed by column index so duplicate header names are preserved.
        header_row_index = first_data_row_index
        if header_row_index and header_row_index <= sheet.max_row:
            for i, col_name in enumerate(truncated_data_values, 1):
                if not col_name:
                    continue
                header_cell = sheet.cell(row=header_row_index, column=i)
                note_text = None
                if header_cell.comment:
                    comment_text = header_cell.comment.text
                    if comment_text and comment_text.strip():
                        note_text = comment_text.strip()
                elif header_row_index > 1:
                    above_cell = sheet.cell(row=header_row_index - 1, column=i)
                    if above_cell.value and str(above_cell.value).strip():
                        note_text = str(above_cell.value).strip()
                if note_text is None and str(col_name).strip().lower() != "lane #":
                    # Row 2 often holds sheet titles; skip for Lane # so we do not treat them as rules.
                    notes_row_index = 2
                    if notes_row_index <= sheet.max_row:
                        note_cell = sheet.cell(row=notes_row_index, column=i)
                        if note_cell.value and str(note_cell.value).strip():
                            note_text = str(note_cell.value).strip()
                if note_text:
                    column_notes_by_excel_col[i] = note_text

        # Check font color to identify black font columns (required columns)
        # Track column indices with black font to handle duplicate column names
        black_font_indices = []  # Store (index, column_name) tuples
        
        for i, value in enumerate(truncated_data_values):
            if i < len(first_data_row):
                cell = first_data_row[i]
                font_color = "black"
                if cell.font and cell.font.color:
                    hex_color = cell.font.color.rgb
                    if hex_color is not None:
                        # Convert to string and handle different formats
                        hex_str = str(hex_color).upper()
                        # Remove 'FF' prefix if present (ARGB format)
                        if hex_str.startswith('FF') and len(hex_str) == 8:
                            hex_str = hex_str[2:]
                        
                        # Check if it's black
                        if hex_str == '000000' or hex_str == '00000000':
                            font_color = "black"
                        else:
                            # Check if it's a shade of grey (R, G, and B are close)
                            try:
                                if len(hex_str) == 6:
                                    r, g, b = int(hex_str[0:2], 16), int(hex_str[2:4], 16), int(hex_str[4:6], 16)
                                    # Check if it's a shade of grey (R, G, and B are close)
                                    if abs(r - g) < 10 and abs(g - b) < 10 and r > 0:  # Grey (not black, not white)
                                        font_color = "grey"
                                    else:
                                        font_color = "other non-black"  # For colors that are not black or grey
                            except (ValueError, IndexError):
                                pass
                
                if font_color == "black":
                    black_font_values.append(value)
                    black_font_indices.append(i)  # Track the index
    
    # Filter the DataFrame to keep only the columns with black font
    # Handle duplicate column names by using positional selection
    if df_rate_card is not None and black_font_indices:
        # Get the original column positions in the dataframe
        # The dataframe columns should correspond to truncated_data_values after setting columns from row
        
        # Check for duplicate column names and keep only black font versions
        seen_columns = {}  # {column_name: index_in_black_font_indices}
        indices_to_keep = []
        
        for idx in black_font_indices:
            col_name = truncated_data_values[idx] if idx < len(truncated_data_values) else None
            if col_name is not None:
                if col_name not in seen_columns:
                    # First occurrence of this column name with black font
                    seen_columns[col_name] = idx
                    indices_to_keep.append(idx)
                else:
                    # Duplicate column name - skip column but merge its notes into the kept column
                    # (the dropped column often holds row-above-header conditional rules).
                    kept_idx = seen_columns[col_name]
                    dup_excel = idx + 1
                    kept_excel = kept_idx + 1
                    n_dup = column_notes_by_excel_col.get(dup_excel)
                    if n_dup:
                        n_kept = column_notes_by_excel_col.get(kept_excel, "")
                        column_notes_by_excel_col[kept_excel] = _merge_column_note(n_kept, n_dup)
                    print(f"   [DEBUG] Duplicate column '{col_name}' found at index {idx}, keeping first occurrence at index {kept_idx}")

        # Select columns by position using iloc
        if indices_to_keep:
            # Map the indices to dataframe column positions
            # The dataframe was truncated to first_column_index columns
            valid_indices = [i for i in indices_to_keep if i < len(df_rate_card.columns)]
            if valid_indices:
                df_filtered_rate_card = df_rate_card.iloc[:, valid_indices]
                # Update black_font_values to match the filtered columns
                black_font_values = [truncated_data_values[i] for i in valid_indices if i < len(truncated_data_values)]
            else:
                df_filtered_rate_card = df_rate_card
        else:
            df_filtered_rate_card = df_rate_card
    else:
        df_filtered_rate_card = df_rate_card

    if valid_indices is None and df_filtered_rate_card is not None:
        valid_indices = list(range(len(df_filtered_rate_card.columns)))

    # Get list of column names
    column_names = df_filtered_rate_card.columns.tolist()

    # One cleaned condition string per output column (aligned with column_names / valid_indices)
    conditions_list = []
    if truncated_data_values is not None and valid_indices is not None:
        for orig_idx in valid_indices:
            excel_i = orig_idx + 1
            raw_condition = column_notes_by_excel_col.get(excel_i, "")
            cleaned_condition = clean_condition_text(raw_condition) if raw_condition else ""
            conditions_list.append(cleaned_condition)
            if cleaned_condition:
                col_nm = truncated_data_values[orig_idx] if orig_idx < len(truncated_data_values) else ""
                print(f"   [DEBUG] Condition for '{col_nm}' (col {excel_i}):")
                print(f"      Raw: {raw_condition[:80]}..." if len(raw_condition) > 80 else f"      Raw: {raw_condition}")
                print(f"      Cleaned: {cleaned_condition[:80]}..." if len(cleaned_condition) > 80 else f"      Cleaned: {cleaned_condition}")

    return df_filtered_rate_card, column_names, conditions_list


def process_business_rules(file_path):
    """
    Process the Business rules tab from a Rate Card Excel file.
    
    Args:
        file_path (str): Path to the file relative to the "input/" folder
    
    Returns:
        dict: Dictionary containing:
            - 'postal_code_zones': list of zone rules with name, country, postal_codes, exclude
            - 'country_regions': list of region rules with name, country, postal_codes, exclude
            - 'no_data_added': list of entries with no data
            - 'raw_rules': all parsed rules as a list of dicts
    """
    # Construct full path from input folder
    input_folder = "input"
    full_path = os.path.join(input_folder, file_path)
    
    # Load the workbook
    workbook = openpyxl.load_workbook(full_path, data_only=True)
    
    # Check if "Business rules" sheet exists
    if "Business rules" not in workbook.sheetnames:
        print(f"   [WARNING] 'Business rules' sheet not found in {file_path}")
        return {
            'postal_code_zones': [],
            'country_regions': [],
            'no_data_added': [],
            'raw_rules': []
        }
    
    sheet = workbook["Business rules"]
    
    # DEBUG: Print sheet info
    print(f"\n{'='*60}")
    print(f"[DEBUG] BUSINESS RULES SHEET ANALYSIS")
    print(f"{'='*60}")
    print(f"   Sheet name: 'Business rules'")
    print(f"   Total rows in sheet: {sheet.max_row}")
    print(f"   Total columns: {sheet.max_column}")
    print(f"   Available sheets: {workbook.sheetnames}")
    
    # STEP 1: Read all rows and filter out empty ones (skip first 2 rows)
    print(f"\n   [DEBUG] Step 1: Reading and filtering rows (skipping first 2 rows)...")
    
    all_rows = []  # Will store (original_row_idx, row_values) tuples
    for row_idx in range(3, sheet.max_row + 1):
        row = sheet[row_idx]
        row_values = [cell.value for cell in row]
        
        # Check if row is empty
        is_empty = all(v is None or (isinstance(v, str) and v.strip() == '') for v in row_values)
        
        if not is_empty:
            all_rows.append((row_idx, row_values))
    
    print(f"   [DEBUG] Total non-empty rows found: {len(all_rows)} (out of {sheet.max_row - 2} after skipping first 2)")
    
    # DEBUG: Print first 20 non-empty rows to see structure
    print(f"\n   [DEBUG] First 20 non-empty rows content:")
    for i, (row_idx, row_values) in enumerate(all_rows[:20]):
        non_empty = [(col_i, v) for col_i, v in enumerate(row_values) if v is not None]
        print(f"      Row {row_idx}: {non_empty}")
    
    if len(all_rows) > 20:
        print(f"      ... and {len(all_rows) - 20} more rows")
    
    # Marker values to look for (case-insensitive)
    markers = ['postal code zones', 'country regions', 'no data added']
    
    # Result structure
    result = {
        'postal_code_zones': [],
        'country_regions': [],
        'no_data_added': [],
        'raw_rules': []
    }
    
    # Track sections and their header columns
    current_section = None
    header_columns = {}  # Maps column index to header name
    waiting_for_header = False  # Flag to indicate we found a marker and are waiting for header row
    
    print(f"\n   [DEBUG] Step 2: Searching for markers: {markers}")
    print(f"   [DEBUG] Structure: MARKER row -> HEADER row (below) -> DATA rows")
    
    # Process non-empty rows
    for i, (row_idx, row_values) in enumerate(all_rows):
        # Check if this row contains a marker (section header)
        row_text_lower = ' '.join(str(v).lower() for v in row_values if v is not None)
        
        found_marker = None
        for marker in markers:
            if marker in row_text_lower:
                found_marker = marker
                print(f"\n   [DEBUG] >>> MARKER FOUND: '{marker}' at row {row_idx}")
                break
        
        if found_marker:
            # This is a marker row - next non-empty row will be the header
            current_section = found_marker.replace(' ', '_')
            waiting_for_header = True
            header_columns = {}  # Reset header columns for new section
            print(f"   [DEBUG]     Section: '{current_section}'")
            print(f"   [DEBUG]     Waiting for header row...")
            continue
        
        # If we're waiting for header, this row should be the header
        if waiting_for_header:
            waiting_for_header = False
            header_columns = {}
            
            print(f"   [DEBUG]     Header row (row {row_idx}): {[v for v in row_values if v is not None]}")
            
            for col_idx, cell_value in enumerate(row_values):
                if cell_value:
                    header_name = str(cell_value).strip().lower()
                    # Normalize header names
                    if 'name' in header_name:
                        header_columns[col_idx] = 'name'
                    elif 'country' in header_name:
                        header_columns[col_idx] = 'country'
                    elif 'postal' in header_name or 'code' in header_name:
                        header_columns[col_idx] = 'postal_code'
                    elif 'exclude' in header_name:
                        header_columns[col_idx] = 'exclude'
                    else:
                        header_columns[col_idx] = header_name
            
            print(f"   [DEBUG]     Mapped header columns: {header_columns}")
            continue
        
        # If we're in a section and have header columns, parse the data row
        if current_section and header_columns:
            rule_data = {
                'section': current_section,
                'name': None,
                'country': None,
                'postal_code': None,
                'exclude': None
            }
            
            # Extract values based on header columns
            for col_idx, header_name in header_columns.items():
                if col_idx < len(row_values):
                    value = row_values[col_idx]
                    if value is not None:
                        rule_data[header_name] = str(value).strip() if value else None
            
            # Only add if we have at least a name or postal code
            if rule_data['name'] or rule_data['postal_code'] or rule_data['country']:
                result['raw_rules'].append(rule_data)
                print(f"   [DEBUG]     + Rule added: {rule_data}")
                
                # Add to appropriate section list
                if current_section == 'postal_code_zones':
                    result['postal_code_zones'].append(rule_data)
                elif current_section == 'country_regions':
                    result['country_regions'].append(rule_data)
                elif current_section == 'no_data_added':
                    result['no_data_added'].append(rule_data)
    
    print(f"\n{'='*60}")
    print(f"[DEBUG] BUSINESS RULES SUMMARY")
    print(f"{'='*60}")
    print(f"   - Postal Code Zones: {len(result['postal_code_zones'])} rules")
    print(f"   - Country Regions: {len(result['country_regions'])} rules")
    print(f"   - No Data Added: {len(result['no_data_added'])} entries")
    print(f"   - Total raw rules: {len(result['raw_rules'])}")
    
    if not result['raw_rules']:
        print(f"\n   [WARNING] No rules were found! Possible issues:")
        print(f"      1. Markers not found in expected format")
        print(f"      2. Headers not in row above markers")
        print(f"      3. Data structure different than expected")
    
    return result


def transform_business_rules_to_conditions(business_rules):
    """
    Transform parsed business rules into condition format.
    
    Args:
        business_rules (dict): Output from process_business_rules()
    
    Returns:
        dict: Dictionary mapping zone/region names to their conditions
              Format: {zone_name: {'country': 'XX', 'postal_codes': ['12', '34'], 'exclude': bool}}
    """
    conditions = {}
    
    for rule in business_rules.get('raw_rules', []):
        name = rule.get('name')
        if not name:
            continue
        
        section = rule.get('section', '')
        
        # Parse postal codes (comma-separated, possibly with spaces)
        # For country_regions, we don't use postal codes - only country matters
        postal_code_str = rule.get('postal_code', '')
        postal_codes = []
        
        if section != 'country_regions' and postal_code_str:
            # Split by comma and clean up each code
            postal_codes = [code.strip() for code in str(postal_code_str).split(',') if code.strip()]
        
        # Excluded column: keep raw value (e.g. "Jiangsu, Jiashan, Jiaxing, Jiangxi") for the result file
        exclude_raw = rule.get('exclude')
        excluded_value = (exclude_raw if exclude_raw is not None else '')
        if not isinstance(excluded_value, str):
            excluded_value = str(excluded_value)
        excluded_value = excluded_value.strip()
        # Boolean for "(EXCLUDE)" in formatted condition: any non-empty value
        is_exclude = bool(excluded_value)

        condition = {
            'section': section,
            'country': rule.get('country'),
            'postal_codes': postal_codes,
            'exclude': is_exclude,
            'excluded_value': excluded_value,  # raw text from Excluded column for result
            'raw_postal_code': postal_code_str if section != 'country_regions' else ''
        }
        
        conditions[name] = condition
    
    return conditions


def format_business_rule_condition(rule_name, condition):
    """
    Format a business rule condition into a readable string.
    
    Args:
        rule_name (str): Name of the rule/zone
        condition (dict): Condition dictionary from transform_business_rules_to_conditions
    
    Returns:
        str: Human-readable condition string
    """
    parts = []
    
    if condition.get('country'):
        parts.append(f"Country: {condition['country']}")
    
    if condition.get('postal_codes'):
        prefix_list = ', '.join(condition['postal_codes'][:5])
        if len(condition['postal_codes']) > 5:
            prefix_list += f", ... (+{len(condition['postal_codes']) - 5} more)"
        parts.append(f"Postal codes starting with: {prefix_list}")
    
    if condition.get('exclude'):
        parts.append("(EXCLUDE)")
    
    return ' | '.join(parts) if parts else 'No conditions'


def find_business_rule_columns(rate_card_df, business_rules_conditions):
    """
    Find which columns in the rate card contain business rule values.
    
    Args:
        rate_card_df (pd.DataFrame): The rate card dataframe
        business_rules_conditions (dict): Dictionary of business rule conditions with rule names as keys
    
    Returns:
        dict: Dictionary with:
            - 'rule_to_columns': {rule_name: [list of columns where found]}
            - 'column_to_rules': {column_name: [list of rules found in it]}
            - 'unique_columns': set of unique column names that contain any business rule
    """
    rule_names = list(business_rules_conditions.keys())
    
    result = {
        'rule_to_columns': {},  # Which columns contain each rule
        'column_to_rules': {},  # Which rules are in each column
        'unique_columns': set()
    }
    
    if rate_card_df is None or rate_card_df.empty or not rule_names:
        return result
    
    print(f"\n{'='*60}")
    print(f"[DEBUG] FINDING BUSINESS RULE COLUMNS IN RATE CARD")
    print(f"{'='*60}")
    print(f"   Searching for {len(rule_names)} rule names in {len(rate_card_df.columns)} columns...")
    
    # Columns to EXCLUDE from business rule detection (these contain codes, not business rule names)
    # Even if a value matches a rule name, these columns should not be treated as business rule columns
    EXCLUDED_BUSINESS_RULE_COLUMNS = {
        'origin airport', 'destination airport', 'origin port', 'destination port',
        'pol', 'poe', 'port of loading', 'port of entry', 'airport', 'port',
        'origin airport code', 'destination airport code', 'airport code',
        'origin seaport', 'destination seaport', 'seaport',
        'ship_port', 'cust_port', 'origin_airport', 'destination_airport',
        'carrier', 'carrier name', 'carrier code', 'scac', 'scac code',
        'origin country', 'destination country', 'country', 'ship_country', 'cust_country'
    }
    
    # Create a set of rule names for faster lookup (case-insensitive)
    rule_names_lower = {str(name).lower(): name for name in rule_names}
    
    # For each column, check which rule names are present
    for col in rate_card_df.columns:
        # Skip excluded columns
        col_lower = str(col).lower().strip()
        if col_lower in EXCLUDED_BUSINESS_RULE_COLUMNS:
            print(f"   [SKIP] Column '{col}' excluded from business rule detection")
            continue
        try:
            # Get unique values in this column
            unique_values = rate_card_df[col].dropna().unique()
            
            # Check each unique value against rule names
            for val in unique_values:
                val_str = str(val).strip().lower()
                
                if val_str in rule_names_lower:
                    original_rule_name = rule_names_lower[val_str]
                    
                    # Track rule to columns mapping
                    if original_rule_name not in result['rule_to_columns']:
                        result['rule_to_columns'][original_rule_name] = []
                    if col not in result['rule_to_columns'][original_rule_name]:
                        result['rule_to_columns'][original_rule_name].append(col)
                    
                    # Track column to rules mapping
                    if col not in result['column_to_rules']:
                        result['column_to_rules'][col] = []
                    if original_rule_name not in result['column_to_rules'][col]:
                        result['column_to_rules'][col].append(original_rule_name)
                    
                    result['unique_columns'].add(col)
        except Exception as e:
            # Skip columns that can't be processed
            pass
    
    # Initialize empty lists for rules not found
    for rule_name in rule_names:
        if rule_name not in result['rule_to_columns']:
            result['rule_to_columns'][rule_name] = []
    
    # Print results
    print(f"\n   [RESULT] Unique columns containing business rules:")
    if result['unique_columns']:
        for col in sorted(result['unique_columns']):
            rules_in_col = result['column_to_rules'].get(col, [])
            print(f"      - '{col}': {len(rules_in_col)} rules found")
            # Show first few rules as examples
            if rules_in_col:
                examples = rules_in_col[:3]
                if len(rules_in_col) > 3:
                    print(f"         Examples: {examples} ... (+{len(rules_in_col) - 3} more)")
                else:
                    print(f"         Rules: {examples}")
    else:
        print(f"      No columns found containing business rule values")
    
    print(f"\n   [SUMMARY] {len(result['unique_columns'])} unique columns contain business rule values")
    
    return result


def get_business_rules_lookup(file_path):
    """
    Get a lookup dictionary from business rule names to their country and postal codes.
    
    Args:
        file_path (str): Path to the rate card file relative to "input/" folder
    
    Returns:
        dict: Dictionary with:
            - 'rule_to_country': {rule_name: country_code}
            - 'rule_to_postal_codes': {rule_name: [list of postal codes]}
            - 'business_rule_columns': set of column names containing business rules
            - 'all_rules': list of all rule data with name, country, postal_codes
    """
    # Process business rules
    business_rules = process_business_rules(file_path)
    business_rules_conditions = transform_business_rules_to_conditions(business_rules)
    
    # Get rate card to find which columns contain business rules
    rate_card_df, rate_card_columns, _ = process_rate_card(file_path)
    business_rule_cols_info = find_business_rule_columns(rate_card_df, business_rules_conditions)
    
    result = {
        'rule_to_country': {},
        'rule_to_postal_codes': {},
        'business_rule_columns': business_rule_cols_info.get('unique_columns', set()),
        'column_to_rules': business_rule_cols_info.get('column_to_rules', {}),
        'all_rules': []
    }
    
    for rule_name, condition in business_rules_conditions.items():
        country = condition.get('country')
        postal_codes = condition.get('postal_codes', [])
        
        if country:
            result['rule_to_country'][rule_name] = country
        if postal_codes:
            result['rule_to_postal_codes'][rule_name] = postal_codes
        
        result['all_rules'].append({
            'name': rule_name,
            'country': country,
            'postal_codes': postal_codes,
            'section': condition.get('section'),
            'exclude': condition.get('exclude', False)
        })
    
    print(f"\n[DEBUG] Business Rules Lookup:")
    print(f"   - Rules with country: {len(result['rule_to_country'])}")
    print(f"   - Rules with postal codes: {len(result['rule_to_postal_codes'])}")
    print(f"   - Columns containing rules: {sorted(result['business_rule_columns'])}")
    
    return result


def get_required_geo_columns():
    """
    Get the list of required geographic columns that should be in the final output.
    These are derived from business rules and should be mapped from ETOF/LC files.
    
    Returns:
        list: List of required column names for origin/destination country and postal codes
    """
    return [
        'Origin Country',
        'Origin Postal Code', 
        'Destination Country',
        'Destination Postal Code'
    ]


def _merge_column_note(existing, new):
    """Merge two condition note strings without duplicating identical blocks."""
    if not new:
        return existing or ""
    if not existing:
        return new
    e, n = existing.strip(), new.strip()
    if n in e:
        return e
    return e + "\n\n" + n


def clean_condition_text(condition_text):
    """
    Clean up condition text for better readability.
    
    Transforms:
        "Conditional rules:
        1. 33321-6422: TOPOSTALCODE starts with 33321-6422,333216422"
    To:
        "1. 33321-6422: starts with 33321-6422,333216422"
    """
    if not condition_text:
        return condition_text
    
    # Remove "Conditional rules:" header (case insensitive)
    cleaned = re.sub(r'(?i)^conditional\s*rules\s*:\s*\n?', '', condition_text.strip())
    
    # Remove column name references like "TOPOSTALCODE ", "FROMPOSTALCODE ", etc.
    # Pattern: After the colon and value identifier, remove uppercase column names followed by space
    # Example: "33321-6422: TOPOSTALCODE starts with" -> "33321-6422: starts with"
    cleaned = re.sub(r':\s*[A-Z_]+\s+(starts with|contains|equals|is empty|does not contain|does not equal)', r': \1', cleaned)
    
    # Also handle cases without numbered format
    cleaned = re.sub(r'^[A-Z_]+\s+(starts with|contains|equals|is empty|does not contain|does not equal)', r'\1', cleaned, flags=re.MULTILINE)
    
    # Clean up extra whitespace and newlines
    lines = [line.strip() for line in cleaned.split('\n') if line.strip()]
    cleaned = '\n'.join(lines)
    
    return cleaned


def _sanitize_for_json(obj):
    """Convert NaN/NaT and non-JSON-serializable values for json.dump."""
    if isinstance(obj, dict):
        return {k: _sanitize_for_json(v) for k, v in obj.items()}
    if isinstance(obj, list):
        return [_sanitize_for_json(x) for x in obj]
    if isinstance(obj, float) and (math.isnan(obj) or math.isinf(obj)):
        return None
    if pd.isna(obj):
        return None
    if hasattr(obj, 'isoformat'):  # datetime, date, time
        return obj.isoformat()
    return obj


def _normalize_excel_cell_text(val):
    """Normalize Excel cell text (newlines, _x000D_, bullets) for parsing."""
    if val is None:
        return ""
    s = str(val).replace("\r\n", "\n").replace("\r", "\n")
    s = s.replace("_x000D_", "\n").replace("\u2022", " ")
    lines = []
    for line in s.split("\n"):
        line = re.sub(r"\s+", " ", line).strip()
        if line:
            lines.append(line)
    return "\n".join(lines)


def _merged_cost_block_text(sheet, row, start_col):
    """Read text from start_col and start_col+1 (merged / split cells)."""
    a = sheet.cell(row=row, column=start_col).value
    b = None
    if start_col + 1 <= sheet.max_column:
        b = sheet.cell(row=row, column=start_col + 1).value
    sa = str(a).strip() if a is not None else ""
    sb = str(b).strip() if b is not None else ""
    if sa and sb and sa != sb:
        return f"{sa}\n{sb}".strip()
    return sa or sb


def _parse_applies_if_text(text):
    t = _normalize_excel_cell_text(text)
    if not t:
        return ""
    if re.match(r"(?i)applies\s*if\s*:", t):
        t = re.sub(r"(?i)^applies\s*if\s*:\s*", "", t).strip()
    elif re.match(r"(?i)applies\s*if\s+", t):
        t = re.sub(r"(?i)^applies\s*if\s+", "", t).strip()
    t = re.sub(r"^\d+\.\s*", "", t, count=1).strip()
    return t


def _parse_rate_by_rule_line(text):
    """Split 'Rate by: ...' and 'Regular rule' / extra lines into Rate_by and Rule."""
    t = _normalize_excel_cell_text(text)
    if not t:
        return "", ""
    body = re.sub(r"(?i)^rate\s*by\s*:\s*", "", t).strip()
    lines = [ln.strip() for ln in body.split("\n") if ln.strip()]
    if not lines:
        return "", ""
    rate_by = lines[0]
    rule = ""
    if len(lines) > 1:
        rest = "\n".join(lines[1:])
        if re.match(r"(?i)regular\s+rule\s*$", lines[1].strip()):
            rule = "Regular rule"
        else:
            rule = rest
    elif re.search(r"(?i);\s*regular\s+rule\s*$", rate_by):
        rate_by = re.sub(r"(?i);\s*regular\s+rule\s*$", "", rate_by).strip()
        rule = "Regular rule"
    elif re.search(r"(?i)\s+regular\s+rule\s*$", rate_by):
        rate_by = re.sub(r"(?i)\s+regular\s+rule\s*$", "", rate_by).strip()
        rule = "Regular rule"
    return rate_by, rule


def _parse_rounding_rule_cell(text):
    t = _normalize_excel_cell_text(text)
    if not t:
        return ""
    if re.match(r"(?i)rounding\s*rule\s*:", t):
        t = re.sub(r"(?i)^rounding\s*rule\s*:\s*", "", t).strip()
    if "\n" in t:
        return "; ".join(line.strip() for line in t.split("\n") if line.strip())
    return t


def _looks_like_metadata_applies_if_line(text: str) -> bool:
    """Excel row that starts with Applies if (with or without colon)."""
    t = _normalize_excel_cell_text(text)
    if not t:
        return False
    first = t.split("\n")[0].strip()
    if re.match(r"(?i)applies\s*if\s*:", first):
        return True
    if re.match(r"(?i)applies\s*if\s+\S", first):
        return True
    return False


def _cell_looks_like_validity_period_line(text: str) -> bool:
    """Excel metadata row like ``Validity period: to 17.08.2025`` or ``from 18.08.2025``."""
    t = _normalize_excel_cell_text(text).lower()
    if not t:
        return False
    return t.startswith("validity period") or t.startswith("validity:")


def _rounding_cell_looks_like_stray_rate_by_block(text: str) -> bool:
    """
    True when the cell under Rounding was actually a merged ``Rate by: …`` (+ ``Regular rule``) line.
    Used to detect one-row vertical misalignment in non-grouped cost blocks.
    """
    t = _normalize_excel_cell_text(text)
    if not t:
        return False
    if re.search(r"(?i)rate\s*by\s*:\s*\S", t):
        return True
    if re.search(r"(?i);\s*regular\s+rule", t):
        return True
    return False


def _looks_like_cost_type_title_row(text: str) -> bool:
    """
    Top row of a cost metadata block: human-readable cost name, not Applies/Rate/Rounding/bracket.
    """
    t = _normalize_excel_cell_text(text)
    if not t or len(t) > 220:
        return False
    first = t.split("\n")[0].strip()
    if re.match(r"(?i)applies\s*if\s*:", first):
        return False
    if re.match(r"(?i)applies\s*if\s+\S", first):
        return False
    if re.match(r"(?i)rate\s*by\s*:", first):
        return False
    if re.match(r"(?i)rounding\s*rule\s*:", first):
        return False
    if _looks_like_weight_bracket_header(first):
        return False
    return True


def _extract_non_grouped_cost_metadata(sheet, header_row: int, currency_col: int):
    """
    Read Cost_type, Applies_if, Rate_by, Rule, Rounding_rule above the lane header.

    Templates that omit the blank spacer row place the cost title at ``header_row - 4``
    instead of ``header_row - 5``. Without detection, Cost_type is empty, the title is
    parsed as Applies_if, and Rate_by/Rounding_rule receive the wrong rows (lane Costs
    then get empty ``Cost Type``).
    """
    r5 = _merged_cost_block_text(sheet, header_row - 5, currency_col)
    r4 = _merged_cost_block_text(sheet, header_row - 4, currency_col)
    r3 = _merged_cost_block_text(sheet, header_row - 3, currency_col)
    r2 = _merged_cost_block_text(sheet, header_row - 2, currency_col)
    r1 = _merged_cost_block_text(sheet, header_row - 1, currency_col)

    cost_type = _normalize_excel_cell_text(r5)
    applies_if = _parse_applies_if_text(r4)
    rate_by, rule = _parse_rate_by_rule_line(r3)
    rounding = _parse_rounding_rule_cell(r2)

    r5n = _normalize_excel_cell_text(r5)
    r4n = _normalize_excel_cell_text(r4)
    r3n = _normalize_excel_cell_text(r3)

    use_compact = False
    if not r5n.strip() and r4n and r3n:
        rb_guess, _ = _parse_rate_by_rule_line(r3n)
        if _looks_like_cost_type_title_row(r4) and (
            _looks_like_metadata_applies_if_line(r3n)
            or (rb_guess and re.match(r"(?i)applies\s*if", rb_guess.strip()))
        ):
            use_compact = True

    if use_compact:
        cost_type = r4n
        applies_if = _parse_applies_if_text(r3)
        rate_by, rule = _parse_rate_by_rule_line(r2)
        bracket_probe = sheet.cell(row=header_row - 1, column=currency_col).value
        if _looks_like_weight_bracket_header(bracket_probe) or _looks_like_weight_bracket_header(
            r1
        ):
            rounding = ""
        else:
            rounding = _parse_rounding_rule_cell(r1)

    return cost_type, applies_if, rate_by, rule, rounding


def _repair_misplaced_non_grouped_cost_definition(cd: dict) -> dict:
    """
    Fix definitions that were emitted with a one-row vertical offset (legacy extract bug).

    Patterns:
    - **Case C**: ``Cost_type`` empty; title in ``Applies_if``; ``Validity period`` (or similar)
      in ``Rate_by``; ``Applies if`` body (and optional ``Cost to prolong``) in ``Rule``;
      true ``Rate by: …; Regular rule`` in ``Rounding_rule``.
    - **Case D**: Like Case C, but ``Rule`` is only ``Cost to prolong: …`` (no ``Applies if`` text).
      Used for **Fuel Surcharge** monthly rows; ``Cost_type`` becomes the title (e.g.
      ``Fuel Surcharge (May 2025)``), ``Rate_by`` becomes ``Weight/chargeable kg``.
    - **Case B**: ``Cost_type`` empty; cost title in ``Applies_if``; ``Rate_by``/``Rule`` empty;
      ``Rounding_rule`` holds ``Rate by: …`` (and often ``; Regular rule``).
    - **Case A**: ``Cost_type`` empty; title in ``Applies_if``; real ``Applies if`` line wrongly
      in ``Rate_by``; ``Rounding_rule`` holds rate/rule text.
    """
    if not isinstance(cd, dict) or cd.get("grouped_cost"):
        return cd
    ct = str(cd.get("Cost_type") or "").strip()
    app = str(cd.get("Applies_if") or "").strip()
    rb = str(cd.get("Rate_by") or "").strip()
    rule = str(cd.get("Rule") or "").strip()
    rr = str(cd.get("Rounding_rule") or "").strip()
    if ct:
        return cd

    # Case C (validity row shifted into Rate_by; applies-if prose in Rule)
    if (
        app
        and rule
        and rr
        and _looks_like_cost_type_title_row(app)
        and _rounding_cell_looks_like_stray_rate_by_block(rr)
    ):
        rule_low = _normalize_excel_cell_text(rule).lower()
        has_applies = "applies if" in rule_low or _looks_like_metadata_applies_if_line(rule)
        rb_ok = (
            not rb
            or _cell_looks_like_validity_period_line(rb)
            or re.match(r"(?i)applies\s*if", rb)
        )
        if has_applies and rb_ok:
            new_rate_by, new_rule = _parse_rate_by_rule_line(rr)
            if new_rate_by or new_rule:
                applies_parts: list[str] = []
                if rb and _cell_looks_like_validity_period_line(rb):
                    applies_parts.append(rb.strip())
                if rule:
                    applies_parts.append(rule.strip())
                out = dict(cd)
                out["Cost_type"] = app
                out["Applies_if"] = "\n".join(applies_parts).strip()
                out["Rate_by"] = new_rate_by
                out["Rule"] = new_rule
                out["Rounding_rule"] = ""
                return out

    # Case B: rate-by + rule landed in Rounding_rule; title in Applies_if
    if app and not rb and not rule and rr and _rounding_cell_looks_like_stray_rate_by_block(rr):
        if _looks_like_cost_type_title_row(app):
            new_rate_by, new_rule = _parse_rate_by_rule_line(rr)
            if new_rate_by or new_rule:
                out = dict(cd)
                out["Cost_type"] = app
                out["Applies_if"] = ""
                out["Rate_by"] = new_rate_by
                out["Rule"] = new_rule
                out["Rounding_rule"] = ""
                return out

    # Case D: Fuel Surcharge (and similar calendar tiers) — same merged cells as Case C, but
    # ``Rule`` is ``Cost to prolong: …`` (no ``Applies if`` paragraph), so Case C never runs.
    if (
        not ct
        and app
        and rb
        and rr
        and _looks_like_cost_type_title_row(app)
        and _cell_looks_like_validity_period_line(rb)
        and _rounding_cell_looks_like_stray_rate_by_block(rr)
    ):
        rule_low = _normalize_excel_cell_text(rule).lower()
        if "applies if" not in rule_low and not _looks_like_metadata_applies_if_line(rule or ""):
            new_rate_by, new_rule = _parse_rate_by_rule_line(rr)
            if new_rate_by:
                out = dict(cd)
                out["Cost_type"] = app
                applies_parts: list[str] = [rb.strip()]
                if rule and str(rule).strip():
                    applies_parts.append(str(rule).strip())
                out["Applies_if"] = "\n".join(applies_parts).strip()
                out["Rate_by"] = new_rate_by
                out["Rule"] = new_rule or ""
                out["Rounding_rule"] = ""
                return out

    # Case A
    if not app or not rb:
        return cd
    if not re.match(r"(?i)applies\s*if", rb):
        return cd
    if not _looks_like_cost_type_title_row(app):
        return cd
    new_rate_by, new_rule = _parse_rate_by_rule_line(rr)
    out = dict(cd)
    out["Cost_type"] = app
    out["Applies_if"] = _parse_applies_if_text(rb)
    out["Rate_by"] = new_rate_by
    out["Rule"] = new_rule
    out["Rounding_rule"] = ""
    return out


def sanitize_filtered_rate_card_json_object(data: dict) -> dict:
    """
    Fix legacy JSON where non-grouped ``cost_definitions`` rows were shifted (empty
    ``Cost_type``, title in ``Applies_if``, etc.) and backfill empty lane ``Cost Type``
    when there is exactly one non-grouped definition.

    Fuel Surcharge calendar rows are normalized by :func:`_repair_misplaced_non_grouped_cost_definition`
    (Case D); lane ``Costs`` titles are then aligned via :func:`_backfill_lane_fuel_surcharge_cost_types`.
    """
    if not isinstance(data, dict):
        return data
    defs = data.get("cost_definitions")
    if isinstance(defs, list):
        fixed = []
        for x in defs:
            if isinstance(x, dict):
                fixed.append(_repair_misplaced_non_grouped_cost_definition(dict(x)))
            else:
                fixed.append(x)
        data["cost_definitions"] = fixed
        defs = fixed
        _backfill_lane_fuel_surcharge_cost_types(data)
    single_ct = ""
    if isinstance(defs, list) and len(defs) == 1:
        d0 = defs[0]
        if isinstance(d0, dict) and not d0.get("grouped_cost"):
            single_ct = str(d0.get("Cost_type") or "").strip()
    if not _refresh_rate_card_costs_from_summary_source(data):
        if single_ct:
            for lane in data.get("rate_card_data") or []:
                if not isinstance(lane, dict):
                    continue
                for c in lane.get("Costs") or []:
                    if isinstance(c, dict) and not str(c.get("Cost Type") or "").strip():
                        c["Cost Type"] = single_ct
    return data


def _backfill_lane_fuel_surcharge_cost_types(data: dict) -> None:
    """
    After Case D repair, each monthly Fuel Surcharge has ``Cost_type`` ``Fuel Surcharge (…)``.
    Lane ``Costs`` rows are still a band of empty ``Cost Type`` cells in file order — copy the
    matching definition title onto each row so lookups work without positional heuristics.
    """
    defs = data.get("cost_definitions") or []
    if not isinstance(defs, list):
        return
    fuel: list[dict] = []
    for d in defs:
        if not isinstance(d, dict) or d.get("grouped_cost"):
            continue
        ct = str(d.get("Cost_type") or "").strip()
        if ct.lower().startswith("fuel surcharge ("):
            fuel.append(d)
    if not fuel:
        return
    for lane in data.get("rate_card_data") or []:
        if not isinstance(lane, dict):
            continue
        costs = lane.get("Costs") or []
        idxs: list[int] = []
        for i, c in enumerate(costs):
            if not isinstance(c, dict):
                continue
            if not str(c.get("Cost Type") or "").strip():
                idxs.append(i)
            elif idxs:
                break
        n = min(len(idxs), len(fuel))
        for j in range(n):
            title = str(fuel[j].get("Cost_type") or "").strip()
            if title:
                costs[idxs[j]]["Cost Type"] = title


def _coerce_cost_price(val):
    """Return int/float when the cell is numeric; otherwise the original value."""
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return None
    if isinstance(val, (int, float)):
        if isinstance(val, float) and val == int(val):
            return int(val)
        return val
    s = str(val).strip().replace(",", "")
    if s == "":
        return None
    try:
        f = float(s)
        if f == int(f):
            return int(f)
        return f
    except ValueError:
        return val


def _json_currency(val):
    """Currency for JSON: blank cells become empty string."""
    if val is None or (isinstance(val, str) and not str(val).strip()):
        return ""
    return val


def _json_price(val):
    """Price for JSON: blank -> \"\"; numbers stay int/float."""
    if val is None or (isinstance(val, str) and not str(val).strip()):
        return ""
    if isinstance(val, float) and pd.isna(val):
        return ""
    coerced = _coerce_cost_price(val)
    return coerced if coerced is not None else ""


def _looks_like_weight_bracket_header(val):
    """True if row-above-header cell looks like '<= 1', '<=150', etc."""
    if val is None:
        return False
    s = str(val).strip().lower()
    if not s:
        return False
    return bool(re.match(r"^<=\s*\d", s))


def _normalize_weight_bracket_label(val):
    """e.g. '<= 1' -> '<=1' for stable output."""
    if val is None:
        return ""
    s = str(val).strip()
    if not s:
        return ""
    return re.sub(r"<=\s*", "<=", s, count=1, flags=re.IGNORECASE)


def _is_grouped_cost_title(text):
    """True if row title is a grouped cost (e.g. 'Grouped cost: Transport cost (...)')."""
    if not text:
        return False
    return "grouped cost:" in str(text).lower()


def _parse_grouped_cost_details_cell(text):
    """
    Parse 'Grouped cost details:' cell into calculation_rule, cost_split, sub_costs list,
    and validity period (from/to + text).
    Returns None if not a grouped-details cell.
    """
    if not text:
        return None
    t = _normalize_excel_cell_text(text)
    if "grouped cost details" not in t.lower():
        return None
    out = {
        "calculation_rule": "",
        "cost_split": "",
        "sub_costs": [],
        "validity_from": "",
        "validity_to": "",
        "validity_text": "",
    }
    m = re.search(r"Calculation\s*rule:\s*([^\n]+)", t, re.I)
    if m:
        out["calculation_rule"] = m.group(1).strip()
    m = re.search(r"Cost\s*split:\s*([^\n]+)", t, re.I)
    if m:
        out["cost_split"] = m.group(1).strip()
    subm = re.search(r"Sub-costs:\s*(.*?)(?=Validity\s*period:|$)", t, re.S | re.I)
    if subm:
        block = subm.group(1)
        for line in block.split("\n"):
            line = re.sub(r"^[•\-\*]\s*", "", line.strip())
            if line:
                out["sub_costs"].append(line)
    m = re.search(r"Validity\s*period:\s*from\s+(\S+)\s+to\s+(\S+)", t, re.I)
    if m:
        out["validity_from"] = m.group(1).strip()
        out["validity_to"] = m.group(2).strip()
        out["validity_text"] = f"from {out['validity_from']} to {out['validity_to']}"
    return out


def find_rate_card_lane_header_row(sheet):
    """
    Find the row that contains lane column headers and at least one 'Currency' cost column.
    Expects column A to be 'Lane #'.
    """
    max_scan = min(sheet.max_row, 200)
    for row_index in range(1, max_scan + 1):
        first = sheet.cell(row=row_index, column=1).value
        if first is None or str(first).strip() != "Lane #":
            continue
        row_values = [sheet.cell(row=row_index, column=c).value for c in range(1, sheet.max_column + 1)]
        if "Currency" in row_values:
            return row_index
    return None


def extract_rate_card_costs_from_sheet(file_path):
    """
    Extract cost-type metadata (rows above the lane header) and per-lane cost values
    from the 'Rate card' sheet. file_path is relative to the input/ folder.

    Grouped cost blocks (title starts with \"Grouped cost:\") expose:
    - grouped_cost_details (calculation_rule, cost_split, sub_costs, validity_*)
    - sub_cost_definitions (per equipment column: sub_cost_name, Rate_by, Rule, Rounding_rule)
    Per-lane lines for those columns use Cost Type = sub-cost name, Grouped under = group title,
    plus validity and calculation fields copied for convenience.

    Returns:
        dict with keys:
            cost_definitions: list of dicts (Cost_type, Applies_if, Rate_by, Rule, Rounding_rule;
                grouped blocks add grouped_cost, grouped_cost_details, sub_cost_definitions)
            costs_by_lane: dict mapping Lane # string -> list of per-cost dicts
    """
    input_folder = "input"
    full_path = os.path.join(input_folder, file_path)
    if not os.path.isfile(full_path):
        return {"cost_definitions": [], "costs_by_lane": {}}

    workbook = openpyxl.load_workbook(full_path, data_only=True)
    if "Rate card" not in workbook.sheetnames:
        return {"cost_definitions": [], "costs_by_lane": {}}

    sheet = workbook["Rate card"]
    header_row = find_rate_card_lane_header_row(sheet)
    if header_row is None:
        return {"cost_definitions": [], "costs_by_lane": {}}

    row_values = [sheet.cell(row=header_row, column=c).value for c in range(1, sheet.max_column + 1)]
    currency_cols = []
    for c, v in enumerate(row_values, start=1):
        if v is not None and str(v).strip() == "Currency":
            currency_cols.append(c)

    if not currency_cols:
        return {"cost_definitions": [], "costs_by_lane": {}}

    bracket_row = header_row - 1  # Row with '<= 1', '<= 2', ... above Flat/p.unit headers

    cost_definitions = []

    for idx, currency_col in enumerate(currency_cols):
        next_curr = currency_cols[idx + 1] if idx + 1 < len(currency_cols) else sheet.max_column + 1
        price_cols_for_def = list(range(currency_col + 1, next_curr))

        t5 = ""
        if header_row > 5:
            t5 = _normalize_excel_cell_text(
                _merged_cost_block_text(sheet, header_row - 5, currency_col)
            )
        t4 = ""
        if header_row > 4:
            t4 = _normalize_excel_cell_text(
                _merged_cost_block_text(sheet, header_row - 4, currency_col)
            )
        cost_type = t5
        if _is_grouped_cost_title(t5):
            cost_type = t5
        elif _is_grouped_cost_title(t4):
            cost_type = t4

        is_grouped = _is_grouped_cost_title(cost_type)

        if is_grouped and header_row > 3:
            details_raw = _merged_cost_block_text(sheet, header_row - 3, currency_col)
            parsed_details = _parse_grouped_cost_details_cell(details_raw)
            sub_cost_definitions = []
            for pcol in price_cols_for_def:
                sn = _normalize_excel_cell_text(sheet.cell(row=header_row - 4, column=pcol).value)
                rate_raw = _merged_cost_block_text(sheet, header_row - 2, pcol)
                rate_by, rule = _parse_rate_by_rule_line(rate_raw)
                round_raw = _merged_cost_block_text(sheet, header_row - 1, pcol)
                rounding = _parse_rounding_rule_cell(round_raw)
                sub_cost_definitions.append({
                    "sub_cost_name": sn,
                    "Rate_by": rate_by,
                    "Rule": rule,
                    "Rounding_rule": rounding,
                })
            cost_definitions.append({
                "Cost_type": cost_type,
                "grouped_cost": True,
                "Applies_if": "",
                "Rate_by": "",
                "Rule": "",
                "Rounding_rule": "",
                "grouped_cost_details": parsed_details,
                "sub_cost_definitions": sub_cost_definitions,
            })
            continue

        cost_type_ng, applies_if, rate_by, rule, rounding = _extract_non_grouped_cost_metadata(
            sheet, header_row, currency_col
        )
        cd_ng = _repair_misplaced_non_grouped_cost_definition(
            {
                "Cost_type": cost_type_ng,
                "grouped_cost": False,
                "Applies_if": applies_if,
                "Rate_by": rate_by,
                "Rule": rule,
                "Rounding_rule": rounding,
            }
        )
        cost_definitions.append(cd_ng)

    costs_by_lane = {}
    for row_index in range(header_row + 1, sheet.max_row + 1):
        lane_cell = sheet.cell(row=row_index, column=1).value
        if lane_cell is None or (isinstance(lane_cell, str) and not str(lane_cell).strip()):
            continue
        lane_key = str(lane_cell).strip()

        lane_costs = []
        for i, currency_col in enumerate(currency_cols):
            next_currency = currency_cols[i + 1] if i + 1 < len(currency_cols) else sheet.max_column + 1
            price_cols = list(range(currency_col + 1, next_currency))
            cd = cost_definitions[i] if i < len(cost_definitions) else {}
            cost_type_name = cd.get("Cost_type", "") if isinstance(cd, dict) else ""
            is_grouped = isinstance(cd, dict) and cd.get("grouped_cost") is True
            parsed_gc = (cd.get("grouped_cost_details") or {}) if isinstance(cd, dict) else {}
            currency_val = sheet.cell(row=row_index, column=currency_col).value

            if not price_cols:
                continue

            # Wide layout: many columns (weight bracket row + Flat/p.unit per column), e.g. DHL
            if len(price_cols) > 1:
                # Grouped cost: one column per sub-cost; row above header has sub-cost names + grouped details
                if is_grouped:
                    for pcol in price_cols:
                        wb = sheet.cell(row=bracket_row, column=pcol).value
                        measurement = sheet.cell(row=header_row, column=pcol).value
                        price_val = sheet.cell(row=row_index, column=pcol).value
                        sub_name = _normalize_excel_cell_text(
                            sheet.cell(row=header_row - 4, column=pcol).value
                        )
                        lane_costs.append({
                            "Cost Type": sub_name or cost_type_name,
                            "Grouped under": cost_type_name,
                            "Calculation rule": parsed_gc.get("calculation_rule", ""),
                            "Cost split": parsed_gc.get("cost_split", ""),
                            "Validity period": parsed_gc.get("validity_text", ""),
                            "Validity from": parsed_gc.get("validity_from", ""),
                            "Validity to": parsed_gc.get("validity_to", ""),
                            "Currency": _json_currency(currency_val),
                            "Measurement": measurement if measurement is not None else "",
                            "Weight Bracket": _normalize_weight_bracket_label(wb),
                            "Price": _json_price(price_val),
                        })
                    continue
                # DHL-style: weight brackets in row above header (e.g. '<= 1')
                use_weight_brackets = _looks_like_weight_bracket_header(
                    sheet.cell(row=bracket_row, column=price_cols[0]).value
                ) or any(
                    _looks_like_weight_bracket_header(sheet.cell(row=bracket_row, column=c).value)
                    for c in price_cols[: min(5, len(price_cols))]
                )
                if use_weight_brackets:
                    for pcol in price_cols:
                        wb = sheet.cell(row=bracket_row, column=pcol).value
                        measurement = sheet.cell(row=header_row, column=pcol).value
                        price_val = sheet.cell(row=row_index, column=pcol).value
                        lane_costs.append({
                            "Cost Type": cost_type_name,
                            "Currency": _json_currency(currency_val),
                            "Measurement": measurement if measurement is not None else "",
                            "Weight Bracket": _normalize_weight_bracket_label(wb),
                            "Price": _json_price(price_val),
                        })
                    continue
                # Wide fallback (no grouped flag, no weight row): repeat global cost type per column
                for pcol in price_cols:
                    wb = sheet.cell(row=bracket_row, column=pcol).value
                    measurement = sheet.cell(row=header_row, column=pcol).value
                    price_val = sheet.cell(row=row_index, column=pcol).value
                    lane_costs.append({
                        "Cost Type": cost_type_name,
                        "Currency": _json_currency(currency_val),
                        "Measurement": measurement if measurement is not None else "",
                        "Weight Bracket": _normalize_weight_bracket_label(wb),
                        "Price": _json_price(price_val),
                    })
                continue

            # Single column after Currency: legacy (KN) or one wide bracket column
            pcol = price_cols[0]
            wb = sheet.cell(row=bracket_row, column=pcol).value
            price_val = sheet.cell(row=row_index, column=pcol).value
            meas_header = sheet.cell(row=header_row, column=pcol).value

            if _looks_like_weight_bracket_header(wb):
                lane_costs.append({
                    "Cost Type": cost_type_name,
                    "Currency": _json_currency(currency_val),
                    "Measurement": meas_header if meas_header is not None else "",
                    "Weight Bracket": _normalize_weight_bracket_label(wb),
                    "Price": _json_price(price_val),
                })
            else:
                weight_bracket = ""
                if currency_col + 2 <= sheet.max_column:
                    wh = sheet.cell(row=header_row, column=currency_col + 2).value
                    if wh is not None and "weight" in str(wh).lower():
                        weight_bracket = sheet.cell(row=row_index, column=currency_col + 2).value
                lane_costs.append({
                    "Cost Type": cost_type_name,
                    "Currency": _json_currency(currency_val),
                    "Measurement": meas_header if meas_header is not None else "",
                    "Weight Bracket": weight_bracket if weight_bracket not in (None, "") else "",
                    "Price": _json_price(price_val),
                })

        costs_by_lane[lane_key] = lane_costs

    return {"cost_definitions": cost_definitions, "costs_by_lane": costs_by_lane}


def _normalize_lane_key(lane_val):
    if lane_val is None or (isinstance(lane_val, float) and pd.isna(lane_val)):
        return ""
    if isinstance(lane_val, float) and lane_val == int(lane_val):
        return str(int(lane_val))
    return str(lane_val).strip()


def _attach_costs_to_rate_card_records(rate_card_records, costs_by_lane):
    """Add 'Costs' list to each lane record using Lane # lookup."""
    for rec in rate_card_records:
        lk = _normalize_lane_key(rec.get("Lane #"))
        costs = costs_by_lane.get(lk)
        if costs is None and lk:
            try:
                costs = costs_by_lane.get(str(int(float(lk))))
            except (ValueError, TypeError):
                costs = None
        rec["Costs"] = list(costs) if costs is not None else []


def _refresh_rate_card_costs_from_summary_source(data: dict) -> bool:
    """
    If ``summary['Source File']`` names an .xlsx under ``input/`` next to this module,
    re-extract ``cost_definitions`` and per-lane ``Costs`` so lane rows get correct
    ``Cost Type`` after metadata repair (multi-column layouts).
    """
    if not isinstance(data, dict):
        return False
    summary = data.get("summary") or {}
    src = summary.get("Source File")
    if not isinstance(src, str) or not src.strip():
        return False
    basename = os.path.basename(src.strip())
    if not basename.lower().endswith(".xlsx"):
        return False
    script_dir = os.path.dirname(os.path.abspath(__file__))
    abs_path = os.path.join(script_dir, "input", basename)
    if not os.path.isfile(abs_path):
        return False
    bundle = extract_rate_card_costs_from_sheet(abs_path)
    cds = bundle.get("cost_definitions")
    costs_by_lane = bundle.get("costs_by_lane") or {}
    if not isinstance(cds, list) or not costs_by_lane:
        return False
    data["cost_definitions"] = cds
    for lane in data.get("rate_card_data") or []:
        if not isinstance(lane, dict):
            continue
        lk = _normalize_lane_key(lane.get("Lane #"))
        fresh = costs_by_lane.get(lk)
        if fresh is None and lk:
            try:
                fresh = costs_by_lane.get(str(int(float(lk))))
            except (ValueError, TypeError):
                fresh = None
        if isinstance(fresh, list):
            lane["Costs"] = [dict(x) for x in fresh]
    return True


def _has_business_rule_for_cell(rule_name_value, column_name, business_rules_data):
    """
    True if there exists a business rule where Rule Name equals rule_name_value
    and Rate Card Columns contains column_name.
    """
    if rule_name_value is None or (isinstance(rule_name_value, str) and not rule_name_value.strip()):
        return False
    val_str = str(rule_name_value).strip()
    col_str = str(column_name).strip()
    for rule in business_rules_data:
        r_name = rule.get('Rule Name')
        r_cols = rule.get('Rate Card Columns', '')
        if r_name is None or pd.isna(r_name):
            continue
        if str(r_name).strip() != val_str:
            continue
        # Rate Card Columns can be "Origin City" or "Destination City, Origin City"
        cols = [c.strip() for c in str(r_cols).split(',') if c and c.strip()]
        if col_str in cols:
            return True
    return False


def _has_conditional_rule_for_cell(cell_value, column_name, conditions_data, column_index=None):
    """
    True if the Condition Rule for this column (by position when column_index is set)
    contains the cell value (substring, case-insensitive).
    """
    if cell_value is None or (isinstance(cell_value, str) and not cell_value.strip()):
        return False
    val_str = str(cell_value).strip().lower()
    col_str = str(column_name).strip()
    if column_index is not None and 0 <= column_index < len(conditions_data):
        cond = conditions_data[column_index]
        if str(cond.get('Column', '')).strip() == col_str:
            rule_text = cond.get('Condition Rule') or ''
            return val_str in str(rule_text).lower()
    for cond in conditions_data:
        c_col = cond.get('Column')
        if c_col is None or str(c_col).strip() != col_str:
            continue
        rule_text = cond.get('Condition Rule') or ''
        if val_str in str(rule_text).lower():
            return True
    return False


def _enrich_rate_card_records_for_json(rate_card_records, column_names, conditions_data, business_rules_data):
    """
    For each record, for each column except "Lane #", add two keys:
    - "{ColumnName} - Has Business Rule": "Yes"/"No"
    - "{ColumnName} - Has conditional Rule": "Yes"/"No"
    """
    EXCLUDED_COLUMN = "Lane #"
    enriched = []
    for record in rate_card_records:
        new_record = []
        for j, col in enumerate(column_names):
            new_record.append((col, record.get(col)))
            if col == EXCLUDED_COLUMN:
                continue
            cell_value = record.get(col)
            has_br = _has_business_rule_for_cell(cell_value, col, business_rules_data)
            has_cr = _has_conditional_rule_for_cell(cell_value, col, conditions_data, column_index=j)
            new_record.append((f"{col} - Has Business Rule", "Yes" if has_br else "No"))
            new_record.append((f"{col} - Has conditional Rule", "Yes" if has_cr else "No"))
        enriched.append(dict(new_record))
    return enriched


def save_rate_card_to_json(rate_card_dataframe, conditions_data, business_rules_data, summary_data,
                           output_path=None, folder_name="partly_df", source_file_path=None):
    """
    Save rate card output to a JSON file from the same in-memory data (not from the Excel file).

    Args:
        rate_card_dataframe (pd.DataFrame): Rate card data
        conditions_data (list): List of dicts with Column, Has Condition, Condition Rule
        business_rules_data (list): List of dicts for Business Rules sheet
        summary_data (dict or list): Summary metrics
        output_path (str): Optional full path. If None, uses
            ``Filtered_Rate_Card_with_Conditions_<RA id>.json`` when ``source_file_path`` is set (RA from filename).
        folder_name (str): Folder under script directory (default: partly_df)
        source_file_path (str): Optional path relative to input/ to read cost columns and derive default JSON name

    Returns:
        str: Path to the saved JSON file
    """
    script_dir = os.path.dirname(os.path.abspath(__file__))
    partly_df_folder = os.path.join(script_dir, folder_name)
    if not os.path.exists(partly_df_folder):
        os.makedirs(partly_df_folder)

    if output_path is None:
        if source_file_path:
            stem = default_filtered_rate_card_basename(source_file_path)
            output_path = os.path.join(partly_df_folder, f"{stem}.json")
        else:
            output_path = os.path.join(partly_df_folder, "Filtered_Rate_Card_with_Conditions.json")
    elif not output_path.lower().endswith('.json'):
        output_path = os.path.join(os.path.dirname(output_path), os.path.splitext(os.path.basename(output_path))[0] + '.json')

    rate_card_records = rate_card_dataframe.to_dict(orient='records')
    column_names = list(rate_card_dataframe.columns)
    rate_card_records = _enrich_rate_card_records_for_json(
        rate_card_records, column_names, conditions_data, business_rules_data
    )

    cost_definitions = []
    if source_file_path:
        cost_bundle = extract_rate_card_costs_from_sheet(source_file_path)
        cost_definitions = cost_bundle.get("cost_definitions", [])
        costs_by_lane = cost_bundle.get("costs_by_lane", {})
        _attach_costs_to_rate_card_records(rate_card_records, costs_by_lane)
    else:
        for rec in rate_card_records:
            rec["Costs"] = []

    if isinstance(summary_data, dict) and 'Metric' in summary_data:
        summary_dict = dict(zip(summary_data['Metric'], summary_data['Value']))
    else:
        summary_dict = summary_data if isinstance(summary_data, dict) else {}
    payload = {
        'cost_definitions': _sanitize_for_json(cost_definitions),
        'rate_card_data': _sanitize_for_json(rate_card_records),
        'conditions': _sanitize_for_json(conditions_data),
        'business_rules': _sanitize_for_json(business_rules_data),
        'summary': _sanitize_for_json(summary_dict)
    }
    payload = sanitize_filtered_rate_card_json_object(payload)

    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(payload, f, indent=2, ensure_ascii=False, default=str)

    print(f"   JSON saved to: {output_path}")
    return output_path


def save_rate_card_output(file_path, output_path=None, save_excel=True, save_json=True):
    """
    Process rate card and save output to Excel and/or JSON.
    
    Args:
        file_path (str): Path to the rate card file relative to "input/" folder
        output_path (str): Optional output path. If None, saves to
            ``partly_df/Filtered_Rate_Card_with_Conditions_<RA id>.xlsx`` (RA from input filename).
        save_excel (bool): If True, write the Excel result file (default True)
        save_json (bool): If True, write the JSON result file (default True)
    
    Returns:
        str: Path to the saved Excel file (or output_path even if only JSON was written)
    """
    # Process the rate card
    rate_card_dataframe, rate_card_column_names, rate_card_conditions = process_rate_card(file_path)
    
    # Process business rules
    business_rules = process_business_rules(file_path)
    business_rules_conditions = transform_business_rules_to_conditions(business_rules)
    
    # Find which columns in rate card contain business rule values
    business_rule_columns = find_business_rule_columns(rate_card_dataframe, business_rules_conditions)
    
    # Set output path - partly_df / Filtered_Rate_Card_with_Conditions_<RA id>.xlsx
    if output_path is None:
        script_dir = os.path.dirname(os.path.abspath(__file__))
        partly_df_folder = os.path.join(script_dir, "partly_df")
        if not os.path.exists(partly_df_folder):
            os.makedirs(partly_df_folder)
        stem = default_filtered_rate_card_basename(file_path)
        output_path = os.path.join(partly_df_folder, f"{stem}.xlsx")
    
    # Create conditions DataFrame with cleaned condition text (one row per column position)
    conditions_data = []
    n_with_condition = 0
    for pos, col_name in enumerate(rate_card_column_names):
        cleaned_condition = ""
        if pos < len(rate_card_conditions):
            cleaned_condition = rate_card_conditions[pos] or ""
        if cleaned_condition:
            n_with_condition += 1
        conditions_data.append({
            'Column': col_name,
            'Column position': pos,
            'Has Condition': 'Yes' if cleaned_condition else 'No',
            'Condition Rule': cleaned_condition
        })
    
    df_conditions = pd.DataFrame(conditions_data)
    
    # Build business_rules_data and summary_data (used for Excel and/or JSON)
    business_rules_data = []
    for rule_name, condition in business_rules_conditions.items():
        rule_columns = business_rule_columns['rule_to_columns'].get(rule_name, [])
        columns_str = ', '.join(rule_columns) if rule_columns else '(not found in data)'
        exclude_display = condition.get('excluded_value', '') or 'No'
        business_rules_data.append({
            'Rule Name': rule_name,
            'Section': condition.get('section', '').replace('_', ' ').title(),
            'Country': condition.get('country', ''),
            'Postal Codes': condition.get('raw_postal_code', ''),
            'Exclude': exclude_display,
            'Rate Card Columns': columns_str,
            'Formatted Condition': format_business_rule_condition(rule_name, condition)
        })
    
    unique_cols_list = sorted(business_rule_columns['unique_columns']) if business_rule_columns['unique_columns'] else ['(none)']
    summary_data = {
        'Metric': [
            'Total Rows',
            'Total Columns',
            'Columns with Conditions',
            'Columns without Conditions',
            'Business Rules - Postal Code Zones',
            'Business Rules - Country Regions',
            'Business Rules - No Data Added',
            'Columns Using Business Rules',
            'Business Rule Column Names',
            'Source File'
        ],
        'Value': [
            len(rate_card_dataframe),
            len(rate_card_column_names),
            n_with_condition,
            len(rate_card_column_names) - n_with_condition,
            len(business_rules.get('postal_code_zones', [])),
            len(business_rules.get('country_regions', [])),
            len(business_rules.get('no_data_added', [])),
            len(business_rule_columns['unique_columns']),
            ', '.join(unique_cols_list),
            file_path
        ]
    }
    
    if save_excel:
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            rate_card_dataframe.to_excel(writer, sheet_name='Rate Card Data', index=False)
            df_conditions.to_excel(writer, sheet_name='Conditions', index=False)
            df_business_rules = pd.DataFrame(business_rules_data)
            if not df_business_rules.empty:
                df_business_rules.to_excel(writer, sheet_name='Business Rules', index=False)
            df_summary = pd.DataFrame(summary_data)
            df_summary.to_excel(writer, sheet_name='Summary', index=False)
        print(f"\n✅ Excel saved to: {output_path}")
        print(f"   - Sheet 'Rate Card Data': {len(rate_card_dataframe)} rows x {len(rate_card_column_names)} columns")
        print(f"   - Sheet 'Conditions': {n_with_condition} columns with conditions")
        print(f"   - Sheet 'Business Rules': {len(business_rules_conditions)} rules")
        print(f"   - Sheet 'Summary': Overview statistics")
    
    if save_json:
        json_path = os.path.join(os.path.dirname(output_path), os.path.splitext(os.path.basename(output_path))[0] + '.json')
        save_rate_card_to_json(
            rate_card_dataframe, conditions_data, business_rules_data, summary_data,
            output_path=json_path, source_file_path=file_path
        )
    
    return output_path


def process_multiple_rate_cards(
    file_paths,
    save_excel=True,
    save_json=True,
    output_path_overrides=None,
):
    """
    Process several rate card workbooks under ``input/``. Each file writes
    ``partly_df/Filtered_Rate_Card_with_Conditions_<RA id>.xlsx`` / ``.json``
    unless ``output_path_overrides`` maps input basename to a custom stem or path.

    Args:
        file_paths: List of paths relative to ``input/`` (e.g. Advanced Export - RA....xlsx)
        output_path_overrides: Optional dict[str, str] mapping ``file_path`` to full output path
            or basename without extension.

    Returns:
        list[str]: Paths to written Excel files (one per input).
    """
    if output_path_overrides is None:
        output_path_overrides = {}
    out_paths = []
    for fp in file_paths:
        override = output_path_overrides.get(fp)
        out_paths.append(
            save_rate_card_output(fp, output_path=override, save_excel=save_excel, save_json=save_json)
        )
    return out_paths


if __name__ == "__main__":
    import sys

    # One or more rate cards under input/ (RA######## in the filename -> Filtered_Rate_Card_with_Conditions_RA########.json)
    _default_inputs = [
        "Advanced Export - RA20250815027 v.11 - SCH GOR25 (Fast Boat) 3.xlsx",
        "Advanced Export - RA20240913026 v.12 - SCH GOR24 (Fast Boat).xlsx"
        
    ]
    INPUT_FILES = sys.argv[1:] if len(sys.argv) > 1 else _default_inputs

    SAVE_EXCEL = False
    SAVE_JSON = True

    process_multiple_rate_cards(INPUT_FILES, save_excel=SAVE_EXCEL, save_json=SAVE_JSON)

    INPUT_FILE = INPUT_FILES[0]
    # Also print to console (first file)
    rate_card_dataframe, rate_card_column_names, rate_card_conditions = process_rate_card(INPUT_FILE)
    print("\nDataFrame shape:", rate_card_dataframe.shape)
    print("\nColumn names:")
    print(rate_card_column_names)
    print("\nConditions (cleaned):")
    for col, condition in zip(rate_card_column_names, rate_card_conditions):
        if not condition:
            continue
        cleaned = condition
        print(f"  {col}: {cleaned[:100]}..." if len(cleaned) > 100 else f"  {col}: {cleaned}")
    
    # Print Business Rules
    print("\n" + "="*60)
    print("BUSINESS RULES")
    print("="*60)
    business_rules = process_business_rules(INPUT_FILE)
    business_rules_conditions = transform_business_rules_to_conditions(business_rules)
    
    print(f"\nParsed {len(business_rules_conditions)} business rules:")
    for rule_name, condition in business_rules_conditions.items():
        formatted = format_business_rule_condition(rule_name, condition)
        print(f"  {rule_name}: {formatted}")
    
    # Find and print which columns contain business rules
    print("\n" + "="*60)
    print("BUSINESS RULE COLUMNS IN RATE CARD")
    print("="*60)
    business_rule_columns = find_business_rule_columns(rate_card_dataframe, business_rules_conditions)
    
    print(f"\nUnique columns containing business rule values:")
    for col in sorted(business_rule_columns['unique_columns']):
        rules_count = len(business_rule_columns['column_to_rules'].get(col, []))
        print(f"  - {col}: {rules_count} rules")
